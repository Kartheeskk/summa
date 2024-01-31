from tkinter import *
from PIL import ImageTk
from tkinter import messagebox,filedialog
import time
import ttkthemes
from tkinter import ttk
import openpyxl
from openpyxl import load_workbook
import re
from tkcalendar import DateEntry
import pandas as pd
import subprocess
import os
#from AppOpener import open,close




#slno auto generate-------------------------------------------------------------------------------------------

def slnofunc():
    file=openpyxl.load_workbook('data1.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_val=sheet.cell(row=row,column=1).value

    try:
        slnos.set(max_row_val+1)
    except:
        slnos.set('1')


def clock():
    global dates,nowtime
    dates=time.strftime('%d/%m/%Y')
    nowtime=time.strftime('%H:%M:%S')
    datetimelabel.config(text=f' Date: {dates}\nTime: {nowtime}')
    datetimelabel.after(1000,clock)




def submit():

    if nameentry.get()=='':
        messagebox.showerror("Error","Name cannot be empty")
    else:
        name=nameentry.get()

    if phoneentry.get() == '':
        phone='-'
    elif len(phoneentry.get()) <10 or len(phoneentry.get()) >10:
        messagebox.showerror("Error","Mobile number should be 10 digits")
    else:
        try:
            phone=int(phoneentry.get())
        except:
            messagebox.showerror("Error","Phone number should no contain characters")

    if altpentry.get() == '':
        alt_phone='-'
    elif len(altpentry.get()) <10 or len(altpentry.get()) >10:
        messagebox.showerror("Error"," Alternnate Mobile number should be 10 digits")
    else:
        try:
            alt_phone=int(altpentry.get())
        except:
            messagebox.showerror("Error","Alternate phone number should no contain characters")


    if aadharentry.get() == '':
        aadhar='-'
    elif len(aadharentry.get()) <12 or len(aadharentry.get()) >12:
        messagebox.showerror("Error","Aadhar should be 12 digits")
    else:
        try:
            aadh=int(aadharentry.get())
            aadhar=str(aadh)
        except:
            messagebox.showerror("Error","Aadhar number should no contain characters")



    if var_a.get()=='Select Stream':
        shift='-'
    else:
        shift=var_a.get()





    if reff.get()=='Reference\n number':
        if refnoentry.get()=='':
            refno='-'
            appli='-'
        elif len(refnoentry.get()) == 6:
            try:
                refno=int(refnoentry.get())
                appli='-'
            except:
                messagebox.showerror("Error","Ref no. should not contain character")
        else:
            messagebox.showerror("Error","Reference number should be 6 digits")


        
    elif reff.get()=='Application\n number':
        if refnoentry.get()=='':
            refno='-'
            appli='-'
        elif len(refnoentry.get()) == 8:
            try:
                appli=int(refnoentry.get())
                refno='-'
            except:
                messagebox.showerror("Error","Application no. should not contain character")
        else:
            messagebox.showerror("Error","Application number should be 8 digits")

    else:
        appli='-'
        refno='-'
        
        


    if remin.get()=='Select' or remin.get()=='Others' :
        if otherentry.get()=='':
            remark='-'
        else:
            remark=otherentry.get()

    else:
        remark=remin.get()


    if reminderentry.get()=='':
        reminder='-'
    else:
        reminder=reminderentry.get()

    
    slno=int(slnos.get())
    date=dates
    time=nowtime
    course=var_b.get()

    

    try:
        file=openpyxl.load_workbook('data1.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=slno)
        sheet.cell(column=2,row=sheet.max_row,value=date)
        sheet.cell(column=3,row=sheet.max_row,value=time)
        sheet.cell(column=4,row=sheet.max_row,value=shift)
        sheet.cell(column=5,row=sheet.max_row,value=course)
        sheet.cell(column=6,row=sheet.max_row,value=name)
        sheet.cell(column=7,row=sheet.max_row,value=refno)
        sheet.cell(column=8,row=sheet.max_row,value=appli)
        sheet.cell(column=9,row=sheet.max_row,value=phone)
        sheet.cell(column=10,row=sheet.max_row,value=alt_phone)
        sheet.cell(column=11,row=sheet.max_row,value=aadhar)
        sheet.cell(column=12,row=sheet.max_row,value=remark)
        sheet.cell(column=13,row=sheet.max_row,value=reminder)
        file.save(r'data1.xlsx')
        messagebox.showinfo("Success","Data Sumitted")


        nameentry.delete(0,END)
        refnoentry.delete(0,END)
        phoneentry.delete(0,END)
        altpentry.delete(0,END)
        aadharentry.delete(0,END)
        reminderentry.delete(0,END)
        otherentry.delete(0,END)
        slnofunc()
        var_a.set('Select Stream')
        remin.set('Select')
        reff.set('Reference\n number')
        
    except:
        messagebox.showerror("Warning","Not Submitter")





    
def clrs():
    res=messagebox.askyesno("Info","Do you want to clear the form")
    if res==True:
        nameentry.delete(0,END)
        refnoentry.delete(0,END)
        phoneentry.delete(0,END)
        altpentry.delete(0,END)
        aadharentry.delete(0,END)
        reminderentry.delete(0,END)
        otherentry.delete(0,END)
        slnofunc()
        var_a.set('Select Stream')
        remin.set('Select')
        reff.set('Reference\n number')
    else:
        pass
    



def exits():
    exmsg=messagebox.askyesno("Warning","Do you want to exit?")
    if exmsg==True:
        root.destroy()
    else:
        pass


def openn():
    #path='data1.xlsx'
    #os.system('data1.xlsx')
    #open('data1.xlsx')
    subprocess.run(['data1.xlsx'],shell=True)








#================================================================================================================================================================
#================================================================================================================================================================

#main--------------------------------------------------------------------------------------------------------------------------------------
    

root=ttkthemes.ThemedTk()
root.get_themes()
root.set_theme('itft1')

root.geometry("{0}x{0}+0+0".format(root.winfo_screenwidth(),root.winfo_screenheight()))
root.configure(bg='#ededed')
root.title('Enquiry Management System')

root.columnconfigure(0,weight=1)
root.columnconfigure(1,weight=1)
root.columnconfigure(2,weight=1)
root.columnconfigure(3,weight=1)
root.columnconfigure(4,weight=1)
root.columnconfigure(5,weight=1)
root.rowconfigure(0,weight=1)
root.rowconfigure(1,weight=1)
root.rowconfigure(2,weight=1)
root.rowconfigure(3,weight=1)
root.rowconfigure(4,weight=1)
root.rowconfigure(5,weight=1)
root.rowconfigure(6,weight=1)
root.rowconfigure(7,weight=1)
#-----------------------------------------------------------------------------------------------------------------------------------------

#left corner------------------------------------------------


datetimelabel=Label(root,font=('times new roman',18,'bold'))
datetimelabel.grid(row=0,column=0,sticky='nw')
clock()

    
titles=Label(root,text="ENQUIRY FORM",font=('times new roman',20,'bold'),bg='#ededed')
titles.grid(row=0,column=2)

#right corner------------------------------------------------

slnos=IntVar()
slnolabel=Label(root,text='Sl.No',font=('times new roman',15,'bold'),bg='#ededed')
slnolabel.grid(row=0,column=4,sticky='ne')
slnoentry=Entry(root,font=('arial',15,'bold'),width=7,bd=3,textvariable=slnos)
slnoentry.grid(row=0,column=5,sticky='nw')
slnofunc()


#sl end---------------------------


#Name Entry----------------------------------------

namelabel=Label(root,text='Name',font=('times new roman',20,'bold'),bg='#ededed')
namelabel.grid(row=1,column=0,sticky='e')
nameentry=Entry(root,font=('arial',15,'bold'),width=24,bd=3)
nameentry.grid(row=1,column=1,sticky='w')


#-----------------

#Shift Entry and course

def fun(*args):
    return var_a.get()
def fun2(*args):
    return var_b.get()


var_a = StringVar()
var_b = StringVar()

courseentry=OptionMenu(root,var_b,'')
courseentry.grid(row=3,column=1,sticky='w')
courseentry.config(width=37,bg='white')

def update_option(*args):
    stream = dicts[var_a.get()]
    var_b.set(stream[0])
    menu = courseentry['menu']
    menu.delete(0,END)
    for i in stream:
        menu.add_command(label=i,command=lambda course=i:var_b.set(course))


shiftlabel=Label(root,text='Stream',font=('times new roman',20,'bold'),bg='#ededed')
shiftlabel.grid(row=2,column=0,sticky='e')
courselabel=Label(root,text='Course',font=('times new roman',20,'bold'),bg='#ededed')
courselabel.grid(row=3,column=0,sticky='e')



dicts={'Select Stream':[' '],
       'GAS':['B.A. Economics','B.A. Defence and Strategic Studies','B.sc. Mathematics','B.sc. Physics','B.sc. Chemistry','B.sc. Plant Biology & PlantBiotechnology',
              'B.sc. Advanced Zoology and Biotechnology','B.com. General','B.com. Corporate Secretaryship','M.A. Economics','M.sc. Chemistry','M.com. General'],
       'SFS I':['B.sc. Visual Communication','B.com. General','B.com. Corporate Secreteryship','B.com. Information System Management','B.A. English','B.sc. Biotechnology',
                'B.sc. Information Technology','B.A. Socialogy','B.sc. Data Analytics','B.com. Computer Application','B.com. Profesional Accounting','B.A. Political science and Administration',
                'M.B.A','M.C.A','M.sc. Mathematics','M.sc. Zoology','M.S.W','M.A. Defence and Stratergic Studies'],
       'SFS II':['B.B.A','B.C.A','B.sc. Computer Science','B.com. General','B.com. Corporate Secreteryship','B.com. Accounting & Finance','B.com. Honours','B.com. Bank management',
                 'B.com. Marketing Management','M.Phil. Economics','M.Phil. Zoology','Ph.D. Economics','Ph.D. Defence and Strategic Studies','Ph.D. Chemistry','Ph.D. Zoology',
                 'Ph.D. Commerce','Ph.D. English','Ph.D. Tamil']
    }




#var_a.trace('w',fun2)
var_a.trace('w',update_option)
var_b.trace('w',fun2)

shiftentry=OptionMenu(root,var_a,*dicts.keys(),command=fun)
shiftentry.grid(row=2,column=1,sticky='w')
shiftentry.config(width=37,bg='white')


var_a.set('Select Stream')

#phone-------------------------------------------------------



phonelabel=Label(root,text='Mobile\nNumber',font=('times new roman',20,'bold'),bg='#ededed')
phonelabel.grid(row=4,column=0,sticky='e')
phoneentry=Entry(root,font=('arial',15,'bold'),width=24,bd=3)
phoneentry.grid(row=4,column=1,sticky='w')


#alt phone------------------------------------------------------

altplabel=Label(root,text='Alternate\nMobile',font=('times new roman',20,'bold'),bg='#ededed')
altplabel.grid(row=5,column=0,sticky='e')
altpentry=Entry(root,font=('arial',15,'bold'),width=24,bd=3)
altpentry.grid(row=5,column=1,sticky='w')


#Right---------------
#-------------------------
#-------------------------
#-------------------------
#-------------------------
#-------------------------
#-------------------------
#-------------------------
#-------------------------
#-------------------------



#Ref no---------------------------------------------------------


#refnolabel=Label(root,text='Ref.no.',font=('times new roman',20,'bold'),bg='#ededed')
reff=StringVar()
refnolabel=OptionMenu(root,reff,'Reference\n number','Application\n number')
refnolabel.config(bg='#ededed',font=('times new roman',14,'bold'),width=9,height=2,bd=1)
refnolabel.grid(row=1,column=2,sticky='e')

reff.set('Reference\n number')
refnoentry=Entry(root,font=('arial',15,'bold'),width=24,bd=3)
refnoentry.grid(row=1,column=3)


#aadhar------------------------------------------------------------


aadharlabel=Label(root,text='Aadhar\nnumber',font=('times new roman',20,'bold'),bg='#ededed')
aadharlabel.grid(row=2,column=2,sticky='e')
aadharentry=Entry(root,font=('arial',15,'bold'),width=24,bd=3)
aadharentry.grid(row=2,column=3)


#remark------------------------------------------------------



remarklabel=Label(root,text='Remarks',font=('times new roman',20,'bold'),bg='#ededed')
remarklabel.grid(row=3,column=2,sticky='e')

remin=StringVar()
remin.set("Select")


remarkentry=OptionMenu(root,remin,'Application Filling Doubts','Details about Courses','Application Fee related Problem','Errors - Submission Time',
                       'Reference no. related Prolems','College Timing Details','Details about Cut-off','College Fee related','Others')
remarkentry.grid(row=3,column=3)

remarkentry.config(width=37,bg='white')










#reminder------------------------------------------------------


reminderlabel=Label(root,text='Reminder\nDate',font=('times new roman',20,'bold'),bg='#ededed')
reminderlabel.grid(row=4,column=2,sticky='e')

reminderentry=DateEntry(root,font=('arial',15,'bold'),width=24,bd=3,selectmode='day',date_pattern='dd-MM-yyyy',firstweekday='sunday')
reminderentry.grid(row=4,column=3)


#other-------------------------------------

otherlabel=Label(root,text='Others',font=('times new roman',20,'bold'),bg='#ededed')
otherlabel.grid(row=5,column=2,sticky='e')
otherentry=Entry(root,font=('arial',15,'bold'),width=24,bd=3)
otherentry.grid(row=5,column=3)



#style-----------------------------------------

mystyle=ttk.Style()
mystyle.configure('success.TButton',font=('Roboto',18))

#Submit--------------------------------------------
img0=PhotoImage(file='Image/submitbtn2.png')
sbm_btn=Button(root,image=img0,command=submit,bd=0,bg='#ededed',cursor='hand2',activebackground='#ededed')
#sbm_btn=ttk.Button(root,text='Submit',style="success.TButton",command=submit)
#sbm_btn.config(height=10,width=20)
sbm_btn.grid(row=6,column=2,sticky='w')

#------------------------------------------------

#img1=PhotoImage(file='resetbtn.png')
#clr_btn=Button(root,image=img1,bd=0,bg='#ededed',cursor='hand2',activebackground='#ededed',command=clrs)
clr_btn=ttk.Button(root,text="Clear Entry",style="success.TButton",command=clrs)
clr_btn.grid(row=1,column=5,sticky='w')



#----------------------------------------------------
img3=PhotoImage(file="Image/exitbt1.png")
exit_btn=Button(root,image=img3,command=exits,bd=0,bg='#ededed',cursor='hand2',activebackground='#ededed')
exit_btn.grid(row=4,column=5,sticky='w')

openbtn=ttk.Button(root,text='Open File',style="success.TButton",command=openn)
openbtn.grid(row=3,column=5,sticky='w')




root.mainloop()
